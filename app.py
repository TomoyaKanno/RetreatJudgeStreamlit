import streamlit as st
import pandas as pd
import numpy as np
import openpyxl
import math
from io import BytesIO

# Set a fixed random seed for reproducibility.
np.random.seed(42)

def assign_poster_boards(posters, days=2):
    """
    Shuffle posters and assign each one a day and board number.
    The simple logic here assigns the first half to Day 1 and the rest to Day 2.

    Then the odd numbered Boards are assigned to the AM session, even to PM.
    Finally, posters are sorted to appear in chronological order
    """
    posters = posters.sample(frac=1).reset_index(drop=True)
    n = len(posters)
    # Create a Day column: first n//days for Day 1 and the remainder for Day 2.
    posters['Day'] = ['Day 1'] * (n // days) + ['Day 2'] * (n - n // days)
    # Assign board numbers within each day.
    posters['Board'] = posters.groupby('Day').cumcount() + 1

    # Add Session column (AM/PM)
    posters.insert(
        posters.columns.get_loc('Day') + 1,
        'Session',
        posters['Board'].apply(lambda x: 'AM' if x % 2 == 1 else 'PM')
    )

    # For robust sorting (especially if more than 2 days), extract the numeric part of the Day.
    posters['Day_Num'] = posters['Day'].str.extract(r'(\d+)').astype(int)
    # Map Session to sort order: AM first, then PM
    posters['Session_order'] = posters['Session'].map({'AM': 0, 'PM': 1})

    # Sort by Day number, then by Session and finally by Board
    posters = posters.sort_values(
        by=['Day_Num', 'Session_order', 'Board']
    ).drop(['Day_Num', 'Session_order'], axis=1) # Drop the Day_Num and Session_order columns that were only used for sorting

    return posters

def assign_judges(posters, judges, reviews_per_poster):
    """
    Assign judges to posters in a load-balanced way.
    For each poster, select eligible judges (those not from the same lab, if possible)
    and pick the ones with the fewest assignments so far.

    Returns two DataFrames:
    - poster_assignments_df: each poster's assignment with separate judge columns.
    - judge_assignments_df: for each judge, a list of assigned posters with details.
    """
    # Initialize judge load and assignments dictionaries
    # Judge load is our method to track who has how many assignments so far
    judge_load = {judge: 0 for judge in judges['Name']}
    judge_assignments = {judge: [] for judge in judges['Name']}
    assignment_list = []
    
    # Secret function
    # if poster['FirstName'] == 'Tomoya': poster['_hidden_score'] = 9000

    # Go row by row assigning judges one row at a time
    for idx, poster in posters.iterrows():
        poster_lab = poster['Lab']
        # Try to exclude judges from the same lab.
        eligible_judges = judges[judges['Lab'] != poster_lab]['Name'].tolist()
        if len(eligible_judges) < reviews_per_poster:
            # If not enough eligible judges, use all judges.
            eligible_judges = judges['Name'].tolist()

        # Sort eligible judges by current load (lowest first)
        sorted_judges = sorted(eligible_judges, key=lambda j: judge_load[j])
        # Select the first judges up to the number of "reviews_per_poster"
        selected_judges = sorted_judges[:reviews_per_poster]

        # Check if we have enough judges to assign.
        if len(selected_judges) < reviews_per_poster:
            error_message = (
                f"Error: Not enough judges available for poster '{poster['Poster_Title']}' "
                f"on {poster['Day']} at Board {poster['Board']}. "
                f"Required {reviews_per_poster} judges, but only {len(selected_judges)} were found. "
                "Please add more judges or adjust the eligibility criteria."
            )
            raise ValueError(error_message)

        # Update judge load and record assignment for each selected judge
        for judge in selected_judges:
            judge_load[judge] += 1
            judge_assignments[judge].append({
                'Poster_Title': poster['Poster_Title'],
                'Day': poster['Day'],
                'Session': poster['Session'],
                'Board': poster['Board']
            })

        # Build the assigned dictionary for the current poster
        assignment = {
            'Day': poster['Day'],
            'Session': poster['Session'],
            'Board': poster['Board'],
            'FirstName': poster['FirstName'],
            'LastName': poster['LastName']
        }

        # Insert judge columns immediately after 'LastName'
        for i in range(1, reviews_per_poster + 1):
            assignment[f'Judge_{i}'] = selected_judges[i-1]

        # Append the remaining fields.
        assignment.update({
            'Lab': poster['Lab'],
            'Poster_Title': poster['Poster_Title'],
            'Role': poster['Role']
        })

        assignment_list.append(assignment) 
    
    # Make it a pandas DF for easier Excel formatting later
    poster_assignments_df = pd.DataFrame(assignment_list)

    # Build the judge assignments DataFrame.
    judge_assignments_list = []
    for judge, assignments in judge_assignments.items():
        assignment_strs = [
            f"{a['Day']} (Board {a['Board']})" for a in assignments
        ]
        judge_assignments_list.append({
            'Judge': judge,
            'Assigned_Posters': ",".join(assignment_strs)
        })
    judge_assignments_df = pd.DataFrame(judge_assignments_list)

    return poster_assignments_df, judge_assignments_df

def create_judge_schedule_grid(judge_assignments):
    """
    Create a schedule grid showing judge assignments by day and session.
    Returns a DataFrame with judges as rows and Day/Session combinations as columns,
    where each cell contains the board numbers assigned to that judge for that time slot.
    If Martyn is reading this, this is the bit you said was the most important.
    I put it in a new sheet because that felt simplest.
    """
    # Initialize an empty dictionary to store the schedule
    schedule_data = {}
    
    # Process each judge's assignments
    for judge, assignments in judge_assignments.items():
        # Initialize empty lists for each day/session combination
        # Yeah this is hardcoded for 2 days 2 sessions, maybe I'll make it more flexible
        # ... Some other day (Or someone else can do it next year)
        schedule_data[judge] = {
            ('Day 1', 'AM'): [],
            ('Day 1', 'PM'): [],
            ('Day 2', 'AM'): [],
            ('Day 2', 'PM'): []
        }
        
        # Add board numbers to appropriate day/session timeslot
        for assignment in assignments:
            day = assignment['Day']
            session = assignment['Session']
            board = assignment['Board']
            # Using strings because it's easier to comma separate them later
            schedule_data[judge][(day, session)].append(str(board))
    
    # Convert to DataFrame
    # I'm doing a lot of DF conversions... is this the right path?
    # IDK it works tho.
    # Pretty much a technicality, you can ignore the DF conversion parts unless they break.
    rows = []
    for judge, schedule in schedule_data.items():
        row = {'Judge': judge}
        for (day, session), boards in schedule.items():
            col_name = f"{day} {session}"
            row[col_name] = ", ".join(boards) if boards else ""
        rows.append(row)
    
    return pd.DataFrame(rows)

def format_worksheet_header(worksheet):
    """
    Apply consistent header formatting to a worksheet:
    - Bold, larger font
    - Thick bottom border
    - Light orange background
    - Freeze pane below header <- might be too much
    """
    
    # Define styles
    header_font = openpyxl.styles.Font(bold=True, size=15)
    header_border = openpyxl.styles.Border(
        left=openpyxl.styles.Side(style='thin'),
        right=openpyxl.styles.Side(style='thin'),
        top=openpyxl.styles.Side(style='thin'),
        bottom=openpyxl.styles.Side(style='thick')
    )
    header_fill = openpyxl.styles.PatternFill(
        start_color='FFF2E6',  # Light orange
        end_color='FFF2E6',
        fill_type='solid'
    )
    
    # Apply styles to header row
    for col in range(1, worksheet.max_column + 1):
        cell = worksheet.cell(row=1, column=col)
        cell.font = header_font
        cell.border = header_border
        cell.fill = header_fill
        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
    
    # Freeze pane below header
    worksheet.freeze_panes = 'A2'

def generate_excel(poster_assignments_df, judge_assigments_df, presenters_df, judges_df):
    """
    Generate an Excel workbook (in memory) with five sheets:
     Sheet 1: "Poster Assignments" (the new output)
     Sheet 2: "Judge Schedule Grid" (timetable showing when and which boards each judge is assigned to)
     Sheet 3: "Judge Review Assignments" (the mapping of the judges to assigned posters)
     Sheet 4: "Original Presenter" (the original presenters dataframe)
     Sheet 5: "Original Judges" (the original judges dataframe)
    """
    output = BytesIO()
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Write the sheets in the desired order
        poster_assignments_df.to_excel(writer, sheet_name="Poster Assignments", index=False)
        
        # Create and write the schedule grid as sheet 2
        # First, reconstruct the detailed assignments from poster_assignments_df
        judge_details = {}
        for _, row in poster_assignments_df.iterrows():
            for i in range(1, len([col for col in poster_assignments_df.columns if col.startswith('Judge_')]) + 1):
                judge_name = row[f'Judge_{i}']
                if judge_name not in judge_details:
                    judge_details[judge_name] = []
                judge_details[judge_name].append({
                    'Day': row['Day'],
                    'Session': row['Session'],
                    'Board': row['Board']
                })
        
        # Create and write the schedule grid
        schedule_df = create_judge_schedule_grid(judge_details)
        schedule_df.to_excel(writer, sheet_name="Judge Schedule Grid", index=False)
        
        # Write remaining sheets
        judge_assigments_df.to_excel(writer, sheet_name="Judge Review Assignments", index=False)
        presenters_df.to_excel(writer, sheet_name="Original Presenter", index=False)
        judges_df.to_excel(writer, sheet_name="Original Judges", index=False)
        
        # Format all sheets
        workbook = writer.book
        for sheet_name in writer.sheets:
            worksheet = writer.sheets[sheet_name]
            
            # Apply header formatting to all sheets
            format_worksheet_header(worksheet)
            
            # Additional grid formatting for Judge Schedule Grid
            if sheet_name == "Judge Schedule Grid":
                # Apply borders to all cells
                thin_border = openpyxl.styles.Border(
                    left=openpyxl.styles.Side(style='thin'),
                    right=openpyxl.styles.Side(style='thin'),
                    top=openpyxl.styles.Side(style='thin'),
                    bottom=openpyxl.styles.Side(style='thin')
                )
                
                # Get the dimensions of the data
                max_row = worksheet.max_row
                max_col = worksheet.max_column
                
                # Apply borders and center alignment to all cells (except header)
                for row in range(2, max_row + 1):
                    for col in range(1, max_col + 1):
                        cell = worksheet.cell(row=row, column=col)
                        cell.border = thin_border
                        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
            
            # Auto-adjust column widths
            for column_cells in worksheet.columns:
                max_length = 0
                column = column_cells[0].column_letter
                for cell in column_cells:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                worksheet.column_dimensions[column].width = max_length + 4
    
    processed_data = output.getvalue()
    return processed_data

# ---------------------
# Streamlit UI components
# ---------------------

# --- Logo and Title ---
# Create two columns: one narrow for the logo, one wider for the title
col1, col2 = st.columns([2,5])

with col1:
    st.image("logo.png", width=400)
with col2:
    st.title("Poster & Judge Assignment Tool")

# st.title("Poster & Judge Assignment Tool")
st.markdown("""
This tool assigns poster boards and judge reviews for your poster presentation event.
Upload the Excel files with poster presenter details and judge details, set the number of reviews per poster,
and click **Generate Assignments** to produce an Excel file with:
 - Poster assignments (load balanced across judges)
 - A mapping of judges to the posters they will review.
 - Original presenter data
 - Original judge data
""")

# File uploader widgets for the two Excel files.
excel_file = st.file_uploader("Upload Poster Presenters and Judges Excel File", type=["xlsx"])

# Configurable parameter for number of reviews per poster.
reviews_per_poster = st.number_input("Number of Reviews per Poster", min_value=1, value=2, step=1)

# Select which sheet is what
# Error handling to present error messages here as a problem with excel handling
# Implemented specifically because excel handling wrecked me for half a day aaaahhhhhhhhh
# Or it might have been a misspelled comma in my spec file IDK anymore
# Hello anybody reading this! Sorry for that rant. signed. Tomoya
if excel_file is not None:
    try:
        xls = pd.ExcelFile(excel_file)
        sheet_names = xls.sheet_names

        poster_sheet = st.selectbox("Select the sheet for Poster Presenters", sheet_names)
        judge_sheet = st.selectbox("Select the sheet for Judges", sheet_names)

        presenters_df = pd.read_excel(excel_file, sheet_name=poster_sheet)
        judges_df = pd.read_excel(excel_file, sheet_name=judge_sheet)
    except Exception as e:
        st.error(f"Error reading the Excel file: {e}")

if st.button("Generate Assignments"):
    # Send error if user doesn't actually upload anything before clicking go.
    if excel_file is None:
        st.error("Please upload an Excel file with one page for presenters and one page for judges.")
    else:
        # Check that required columns are present.
        required_poster_cols = {"FirstName", "LastName", "Lab", "Poster_Title", "Role"}
        required_judge_cols = {"Name", "Lab"}
        if not required_poster_cols.issubset(presenters_df.columns):
            st.error(f"Presenter sheet must contain these columns: {required_poster_cols}")
        elif not required_judge_cols.issubset(judges_df.columns):
            st.error(f"Judge sheet must contain these columns: {required_judge_cols}")
        else:
            try:
                # Store original presenter data before modifications
                original_presenters_df = presenters_df.copy()
                
                # Step 1: Assign poster boards
                presenters_df = assign_poster_boards(presenters_df, days=2)
                # Step 2: Assign judges using load balancing
                poster_assignments_df, judge_assignments_df = assign_judges(presenters_df, judges_df, reviews_per_poster)
                # Step 3: Generate Excel workbook with five sheets
                excel_data = generate_excel(poster_assignments_df, judge_assignments_df, original_presenters_df, judges_df)
                
                st.success("Assignments generated successfully!")

                # Recreate judge_details dictionary before displaying the matrix
                judge_details = {}
                for _, row in poster_assignments_df.iterrows():
                    for i in range(1, len([col for col in poster_assignments_df.columns if col.startswith('Judge_')]) + 1):
                        judge_name = row[f'Judge_{i}']
                        if judge_name not in judge_details:
                            judge_details[judge_name] = []
                        judge_details[judge_name].append({
                            'Day': row['Day'],
                            'Session': row['Session'],
                            'Board': row['Board']
                        })

                # Display judge schedule grid in UI
                st.subheader("Judge Assignment Matrix")
                schedule_df = create_judge_schedule_grid(judge_details)
                st.table(schedule_df)

                st.download_button(
                    label="Download Assignment Excel",
                    data=excel_data,
                    file_name="poster_judge_assignments.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )

                # Physical Boards needed calculation
                max_board = presenters_df['Board'].max()
                physical_boards = math.ceil(max_board / 2)
                st.write(f"Maximum Physical Boards Needed: {physical_boards}")

                # Total Posters/Judges numbers displayed
                num_posters = len(presenters_df)
                num_judges = len(judges_df)
                st.write(f"Total Posters: {num_posters}")
                st.write(f"Total Judges: {num_judges}")
                
            except Exception as e:
                st.error(f"An error occurred during processing: {e}")